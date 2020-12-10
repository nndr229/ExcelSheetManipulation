import openpyxl as xl
import re

def grid_card(type,row_start,row_end,col_1,col_2,col_3,node_value,file,worksheet):
    for row in range(row_start, row_end): #worksheet.max_row
        x = worksheet.cell(row,col_1).value
        y = worksheet.cell(row,col_2).value
        z = worksheet.cell(row,col_3).value

        grid_string= f"GRID,{node_value},0,{x},{y},{z}"

        if (row == row_start):
            file.write(f"{type} for {worksheet.title}")
            file.write("\n")
        file.write(grid_string)
        file.write("\n")
        node_value +=1        
    
    file.write("\n")
    return node_value    

def force_card(type,row_start,row_end,col_1,col_2,col_3,node_value,file,worksheet):
    for row in range(row_start, row_end): #worksheet.max_row
        Fx = worksheet.cell(row,col_1).value
        Fy = worksheet.cell(row,col_2).value
        Fz = worksheet.cell(row,col_3).value
        force_id = re.compile(r"[^LC0].*")
        force_id = force_id.findall(worksheet.title)[0]
        force_string = f"FORCE,{force_id},{node_value},0,1.0,{Fx},{Fy},{Fz}"

        if (row == row_start):
            file.write(f"{type} for {worksheet.title}")
            file.write("\n")
        file.write(force_string)
        file.write("\n")
        node_value += 1

    file.write("\n")
    return node_value

def moment_card(type,row_start,row_end,col_1,col_2,col_3,node_value,file,worksheet):
    for row in range(row_start,row_end):
        # MOMENT,MomentID,Node,0,1.0,Mx,My,Mz
        Mx = worksheet.cell(row,col_1).value
        My = worksheet.cell(row,col_2).value
        Mz = worksheet.cell(row,col_3).value
        moment_id = re.compile(r"[^LC0].*")
        moment_id = moment_id.findall(worksheet.title)[0]

        moment_string = f"MOMENT,{moment_id},{node_value},0,1.0,{Mx},{My},{Mz}"

        if (row == row_start):
            file.write(f"{type} for {worksheet.title}")
            file.write("\n")
        file.write(moment_string)
        file.write("\n")
        node_value +=1    
    file.write("\n")
    return node_value

def create_cards(filename):
    wb = xl.load_workbook(filename, data_only=True)
    sheets_names = wb.sheetnames

    sheets_names = [x for x in sheets_names if re.match(r"LC[0-9]+", x) ]

    for worksheets in wb.sheetnames:
        worksheet = wb[worksheets]
        if worksheet.title in sheets_names:
            text_file = open(f"Output2/{worksheet.title}.txt", "a+")
            LC_id = re.compile(r"[^LC0].*")
            LC_id = LC_id.findall(worksheet.title)[0]
            LC_id = 100 * int(LC_id)
             
            #GRID INERTIAL 
            BH_node_grid_inertial = 3800001 +LC_id
            BH_node_grid_inertial = grid_card("BULKHEAD__GRID__INERTIAL",36,42,2,3,4,BH_node_grid_inertial,text_file,worksheet)          
            
            #FORCE_INERTIAL
            BH_node_f_inertial = 3800001 + LC_id
            BH_node_f_inertial = force_card("BULKHEAD__FORCE__INERTIAL",36,42,5,6,7,BH_node_f_inertial,text_file,worksheet)

            #GRID AERO 
            BH_node_grid_aero = BH_node_grid_inertial
            BH_node_grid_aero = grid_card("BULKHEAD__GRID__AERO",36,42,2,8,9,BH_node_grid_aero,text_file,worksheet)  
           
            # FORCE_AERO
            BH_node_f_aero = BH_node_f_inertial 
            BH_node_f_aero = force_card("BULKHEAD__FORCE__AERO",36,42,10,11,12,BH_node_f_aero,text_file,worksheet)
            
            # HT_VT_GRID_INERTIAL
            ht_vt_grid_node_inertial = 3900001 + LC_id
            ht_vt_grid_node_inertial = grid_card("HT-VT__GRID_INERTIAL",7,10,15,16,17,ht_vt_grid_node_inertial,text_file,worksheet) 
           
            # HT_VT_FORCE_INERTIAL
            ht_vt_force_node_inertial = 3900001 + LC_id
            ht_vt_force_node_inertial = force_card("HT-VT__FORCE__INERTIAL",7,10,18,19,20,ht_vt_force_node_inertial,text_file,worksheet)
     
            # HT_VT_MOMENT_INERTIAL
            ht_vt_mom_node_inertial = 3900001 + LC_id
            ht_vt_mom_node_inertial = moment_card("HT-VT__MOMENT__INERTIAL",7,10,21,22,23,ht_vt_mom_node_inertial,text_file,worksheet)
    
            # HT_VT_GRID_AERO
            ht_vt_grid_node_aero = ht_vt_grid_node_inertial 
            ht_vt_grid_node_aero = grid_card("HT-VT__GRID__AERO",7,10,24,25,26,ht_vt_grid_node_aero,text_file,worksheet)
            
            # HT_VT_FORCE_INERTIAL
            ht_vt_force_node_aero = ht_vt_force_node_inertial
            ht_vt_force_node_aero = force_card("HT-VT__FORCE__AERO",7,10,27,28,29,ht_vt_force_node_aero,text_file,worksheet)
       
            # VF 
            VF_node_grid_i = 3950001 + LC_id
            VF_node_grid_i = grid_card("VF__GRID__INERTIAL",19,20,15,16,17,VF_node_grid_i,text_file,worksheet)
            
            VF_node_force_i = 3950001 + LC_id
            VF_node_force_i = force_card("VF__FORCE__INERTIAL",19,20,18,19,20,VF_node_force_i,text_file,worksheet)

            VF_node_moment_i = 3950001 + LC_id
            VF_node_moment_i = moment_card("VF__MOMENT__INERTIAL",19,20,21,22,23,VF_node_moment_i,text_file,worksheet)

            VF_node_grid_aero = VF_node_grid_i
            VF_node_grid_aero =  grid_card("VF__GRID__AERO",19,20,24,25,26,VF_node_grid_aero,text_file,worksheet)

            VF_node_force_aero = VF_node_force_i
            VF_node_force_aero = force_card("VF__FORCE__INERTIAL",19,20,27,28,29,VF_node_force_aero,text_file,worksheet)


            text_file.close()



create_cards(r"input.xlsx")
