import openpyxl as xl
import re

def force_card(col_1,x,y,z,file,worksheet):
    lc_number = None
    for row in range(1, worksheet.max_row): 
        if worksheet.cell(row,col_1).value and str(worksheet.cell(row,col_1).value) != "GRID":
            if len(re.findall(r"LC[0-9]+",str(worksheet.cell(row,col_1).value)))>0:
                if worksheet.cell(row,col_1).value == re.findall(r"LC[0-9]+",worksheet.cell(row,col_1).value)[0]:
                    file.write("\n")
                    lc_number = worksheet.cell(row,col_1).value
                    file.write(f"$HT-VT Reactions for {worksheet.cell(row,col_1).value}")
                    file.write("\n")
            else:
                Fx = worksheet.cell(row,x).value * -1.0
                Fy = worksheet.cell(row,y).value * -1.0
                Fz = worksheet.cell(row,z).value * -1.0
                grid = worksheet.cell(row,col_1).value
                force_id = re.compile(r"[^LC0].*")
                force_id = force_id.findall(lc_number)[0]
                force_string = f"FORCE,{force_id},{grid},0,1.0,{Fx},{Fy},{Fz}"
                file.write(force_string)
                file.write("\n")     
   

def create_cards(filename):
    wb = xl.load_workbook(filename, data_only=True)
    for worksheets in wb.sheetnames:
        worksheet = wb[worksheets]
        text_file = open(f"ForceCards/HT-VT_Reactions.txt", "a+")   
        force_card(3,4,5,6,text_file,worksheet)
        text_file.close()

create_cards(r"HTVTReactions.xlsx")
